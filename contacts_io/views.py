# contacts_io/views.py
from django.http import HttpResponse, HttpResponseNotAllowed
from django.shortcuts import render
from integration_utils.bitrix24.bitrix_user_auth.main_auth import main_auth

from .forms import UploadForm, ExportForm
from .services.importer import import_contacts_rows
from .services.exporter import export_contacts

import csv
from io import StringIO, BytesIO
from openpyxl import load_workbook, Workbook


# ---------- общий помощник чтения файла для импорта ----------
def _read_rows(fmt, django_file):
    rows, headers = [], []
    if fmt == "csv":
        raw = django_file.read().decode("utf-8-sig")
        reader = csv.DictReader(StringIO(raw))
        rows = list(reader)
        headers = reader.fieldnames or []
    elif fmt == "xlsx":
        wb = load_workbook(filename=BytesIO(django_file.read()), read_only=True)
        ws = wb.active
        first = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [(c.value.strip() if isinstance(c.value, str) else (c.value or "")) for c in first]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, [str(v).strip() if v else "" for v in row])))
    return rows, headers


# ----------------------------- ИМПОРТ -----------------------------
#@main_auth(on_start=True)   # AUTH_ID/DOMAIN прилетают из формы (см. upload.html)
def upload_view(request):
    context = {"form": UploadForm()}
    if request.method == "POST":
        # Авторизация только на POST
        from integration_utils.bitrix24.bitrix_user_auth.main_auth import main_auth
        # Обернём только обработку POST
        def do_post(request):
            form = UploadForm(request.POST, request.FILES)
            action = request.POST.get("action", "preview")
            if form.is_valid():
                fmt = form.cleaned_data["fmt"]
                file = request.FILES["file"]
                rows, headers = _read_rows(fmt, file)
                preview_table = [[r.get(h, "") for h in headers] for r in rows[:5]]
                context.update({
                    "form": UploadForm(),
                    "total": len(rows),
                    "headers": headers,
                    "preview": preview_table,
                })
                if action == "import" and rows:
                    token = getattr(request, "bitrix_user_token", None)
                    if not token:
                        context["auth_error"] = "Не получен AUTH_ID. Откройте страницу из Битрикс24."
                    else:
                        stats = import_contacts_rows(rows, token)
                        context["import_stats"] = stats
            else:
                context["form"] = form
            return render(request, "contacts_io/upload.html", context)
        # main_auth только для POST
        return main_auth(on_start=True)(do_post)(request)
    return render(request, "contacts_io/upload.html", context)


# ----------------------------- ЭКСПОРТ -----------------------------
def export_page(request):
    """
    GET-страница с формой экспорта (без авторизации),
    чтобы не ловить 403 из-за отсутствия AUTH_ID на GET.
    """
    return render(request, "contacts_io/export.html", {"form": ExportForm()})


@main_auth(on_start=True)   # POST — авторизация через AUTH_ID/DOMAIN, которые кладёт JS в форме
def export_download(request):
    print("[DEBUG] Экспорт: старт")
    if request.method != "POST":
        print("[DEBUG] Экспорт: не POST")
        return HttpResponseNotAllowed(["POST"])

    form = ExportForm(request.POST)
    print(f"[DEBUG] Экспорт: данные формы: {request.POST}")
    if not form.is_valid():
        print(f"[DEBUG] Экспорт: невалидная форма: {form.errors}")
        return HttpResponse("Invalid form", status=400)

    fmt = form.cleaned_data["fmt"]
    period = form.cleaned_data["period"]
    print(f"[DEBUG] Экспорт: fmt={fmt}, period={period}")

    token = getattr(request, "bitrix_user_token", None)
    if not token:
        print("[DEBUG] Экспорт: нет токена bitrix_user_token")
        return HttpResponse("Unauthorized (no AUTH_ID)", status=401)

    print("[DEBUG] Экспорт: получаем строки экспорта...")
    try:
        rows = export_contacts(token, period=period)
        print(f"[DEBUG] Экспорт: строк получено: {len(rows)}")
    except Exception as e:
        print(f"[DEBUG] Экспорт: ошибка при получении строк: {e}")
        import traceback; traceback.print_exc()
        return HttpResponse(f"Export error: {e}", status=500)

    headers = ["имя", "фамилия", "номер телефона", "почта", "компания"]
    print(f"[DEBUG] Экспорт: headers: {headers}")

    if fmt == "csv":
        print("[DEBUG] Экспорт: формируем CSV")
        sio = StringIO()
        writer = csv.DictWriter(sio, fieldnames=headers)
        writer.writeheader()
        for r in rows:
            writer.writerow({h: r.get(h, "") for h in headers})
        data = sio.getvalue().encode("utf-8-sig")
        resp = HttpResponse(data, content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = 'attachment; filename="contacts_export.csv"'
        print("[DEBUG] Экспорт: CSV готов, отправляем файл")
        return resp

    # xlsx
    print("[DEBUG] Экспорт: формируем XLSX")
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="contacts_export.xlsx"'
    print("[DEBUG] Экспорт: XLSX готов, отправляем файл")
    return resp
